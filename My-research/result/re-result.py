#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
重构版 ✅ 完全对齐需求
✅ 固定布局：三组数据同表，严格固定行位置
✅ 行维度：指标+任务，列维度：模型
✅ 双功能：总表（10次重复均值±标准差）+ 10次单次实验独立表
✅ 逐行横向高亮（含Avg.行）
✅ 缺失数据兼容，空值不参与比较
"""

import os
import json
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ======================== 配置 ========================
ROOT_DIR = r"D:\user\Documents\ai\paper\1_process\outputs\result"
# ROOT_DIR = r"D:\user\Documents\ai\paper\1_process\outputs\result\xiaorong-new"
OUTPUT_PARENT = r"D:\user\Documents\ai\paper\1_process\outputs\result\！handle"

MODEL_LIST = ["erm", "mldg", "vrex", "dfdn", "darm", "sdagn", "dpjdg", "masfd", "mcpdgE8"]
# MODEL_LIST = ["mcpdgE1", "mcpdgE2", "mcpdgE3", "mcpdgE4", "mcpdgE5", "mcpdgE6", "mcpdgE7", "mcpdgE8"]
TARGET_DATASETS = ["phm", "pu"]
TARGET_TASKS = {
    "phm": ["T1", "T2", "T3", "T4"],
    "pu": ["T5", "T6", "T7", "T8"]
}
TARGET_DATA_SIZES = [
    "1",
    "5",
    "10"
]
# 固定指标顺序（和显示名一一对应）
METRIC_ORDER = [
    "acc",
    "precision_macro",
    "recall_macro",
    "f1_macro"
]
METRIC_DISPLAY = ["Acc", "Pre", "Rec", "F1"]
# 每组shot对应的起始行（严格固定）
SHOT_START_ROW = {
    "1": 1,
    "5": 23,
    "10": 45
}

# 高亮样式
HIGHLIGHT_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
# 列宽设置
COL_WIDTH_EVA = 8
COL_WIDTH_TASK = 8
COL_WIDTH_MODEL = 12

# ======================================================

def get_cell_numeric_value(cell_text):
    """提取单元格数值用于比较，空值/无效值返回None"""
    if cell_text == "-" or cell_text is None or cell_text == "":
        return None
    try:
        # 兼容均值±标准差格式，只取前面的均值
        return float(cell_text.split("±")[0])
    except:
        return None

def load_single_json(json_path):
    """读取单个JSON指标文件，返回×100后的指标值"""
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {k: round(data.get(k, 0) * 100, 4) for k in METRIC_ORDER}
    except Exception as e:
        return None

def load_all_experiment_data():
    """
    全量读取所有实验数据
    返回结构：
    {
        "dataset": {
            "size": {
                "task": {
                    "model": [
                        {"acc": ..., "precision_macro": ...}, # 第1次重复
                        {"acc": ..., "precision_macro": ...}, # 第2次重复
                        ...
                    ]
                }
            }
        }
    }
    """
    full_data = {
        ds: {
            sz: {tk: {m: [] for m in MODEL_LIST} for tk in TARGET_TASKS[ds]}
            for sz in TARGET_DATA_SIZES
        }
        for ds in TARGET_DATASETS
    }

    for model in MODEL_LIST:
        model_path = os.path.join(ROOT_DIR, model)
        if not os.path.exists(model_path):
            continue

        for folder in os.listdir(model_path):
            if "-" not in folder:
                continue
            base_part, repeat_str = folder.rsplit("-", 1)
            if not repeat_str.isdigit():
                continue
            repeat_idx = int(repeat_str) - 1  # 转换为 0-9 的索引

            parts = base_part.split("_")
            if len(parts) != 4:
                continue
            _, dataset, task, size = parts

            if dataset not in TARGET_DATASETS or task not in TARGET_TASKS[dataset] or size not in TARGET_DATA_SIZES:
                continue

            json_file = os.path.join(model_path, folder, "final_test_metrics.json")
            if os.path.exists(json_file):
                result = load_single_json(json_file)
                if result:
                    # 确保列表长度足够，按索引插入
                    while len(full_data[dataset][size][task][model]) <= repeat_idx:
                        full_data[dataset][size][task][model].append(None)
                    full_data[dataset][size][task][model][repeat_idx] = result

    print("✅ 所有实验数据已全部读取完成")
    return full_data

def calc_repeat_stats(metric_list):
    """计算10次重复的均值±标准差，无有效数据返回"-" """
    values = [m for m in metric_list if m is not None]
    if not values:
        return "-"
    mean = np.mean(values)
    std = np.std(values)
    return f"{mean:.2f}±{std:.2f}"

def fill_worksheet_content(ws, dataset, dataset_data, is_summary=True, repeat_idx=None):
    """
    通用工作表内容填充函数
    :param ws: openpyxl工作表对象
    :param dataset: 数据集名称 phm/pu
    :param dataset_data: 该数据集的全量数据
    :param is_summary: 是否为总表（均值±标准差），False为单次实验表
    :param repeat_idx: 单次实验的重复索引（0-9），is_summary=False时必填
    """
    task_list = TARGET_TASKS[dataset]
    model_count = len(MODEL_LIST)

    # 遍历三组shot数据，按固定起始行填充
    for size in TARGET_DATA_SIZES:
        start_row = SHOT_START_ROW[size]
        # 1. 填充表头
        header = ["Eva.", "Task"] + MODEL_LIST
        for col_idx, header_text in enumerate(header, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header_text)
            # 表头加粗可自行添加，这里保留原有样式
        # 2. 填充20行数据
        current_data_row = start_row + 1
        # 遍历4个指标
        for metric_idx, metric_key in enumerate(METRIC_ORDER):
            metric_display = METRIC_DISPLAY[metric_idx]
            # 先存储该指标下4个任务的所有模型数据，用于计算Avg.
            task_model_values = {m: [] for m in MODEL_LIST}
            # 遍历4个任务，填充4行数据
            for task in task_list:
                # 填充Eva.和Task列
                ws.cell(row=current_data_row, column=1, value=metric_display)
                ws.cell(row=current_data_row, column=2, value=task)
                # 遍历所有模型，填充数值
                row_values = []
                for col_idx, model in enumerate(MODEL_LIST, 3):
                    model_repeat_data = dataset_data[size][task][model]
                    cell_val = "-"
                    if is_summary:
                        # 总表：计算10次重复的均值±标准差
                        metric_repeat_list = [d[metric_key] if d is not None else None for d in model_repeat_data]
                        cell_val = calc_repeat_stats(metric_repeat_list)
                        # 提取均值用于Avg.计算
                        numeric_val = get_cell_numeric_value(cell_val)
                        if numeric_val is not None:
                            task_model_values[model].append(numeric_val)
                    else:
                        # 单次表：取对应重复次数的原始数据
                        if repeat_idx < len(model_repeat_data) and model_repeat_data[repeat_idx] is not None:
                            raw_val = model_repeat_data[repeat_idx][metric_key]
                            cell_val = f"{raw_val:.2f}"
                            task_model_values[model].append(raw_val)
                    # 填充单元格
                    ws.cell(row=current_data_row, column=col_idx, value=cell_val)
                    row_values.append(get_cell_numeric_value(cell_val))
                # 本行填充完成，执行高亮
                valid_values = [v for v in row_values if v is not None]
                if valid_values:
                    max_val = max(valid_values)
                    for col_offset, val in enumerate(row_values):
                        if val == max_val:
                            ws.cell(row=current_data_row, column=3 + col_offset).fill = HIGHLIGHT_FILL
                # 下一行
                current_data_row += 1
            # 4个任务填充完成，填充Avg.行
            ws.cell(row=current_data_row, column=1, value=metric_display)
            ws.cell(row=current_data_row, column=2, value="Avg.")
            # 遍历所有模型，计算Avg.
            avg_row_values = []
            for col_idx, model in enumerate(MODEL_LIST, 3):
                model_task_vals = task_model_values[model]
                avg_val = "-"
                if len(model_task_vals) == 4:  # 4个任务都有有效数据
                    avg_numeric = np.mean(model_task_vals)
                    avg_val = f"{avg_numeric:.2f}"
                # 填充单元格
                ws.cell(row=current_data_row, column=col_idx, value=avg_val)
                avg_row_values.append(get_cell_numeric_value(avg_val))
            # Avg.行高亮
            valid_avg_values = [v for v in avg_row_values if v is not None]
            if valid_avg_values:
                max_avg_val = max(valid_avg_values)
                for col_offset, val in enumerate(avg_row_values):
                    if val == max_avg_val:
                        ws.cell(row=current_data_row, column=3 + col_offset).fill = HIGHLIGHT_FILL
            # 下一行
            current_data_row += 1

    # 统一设置列宽
    ws.column_dimensions[get_column_letter(1)].width = COL_WIDTH_EVA
    ws.column_dimensions[get_column_letter(2)].width = COL_WIDTH_TASK
    for col_idx in range(3, 3 + model_count):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTH_MODEL

def generate_dataset_summary_excel(dataset, dataset_data, output_dir):
    """生成数据集总表（10次重复的均值±标准差）"""
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, f"{dataset}_result.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "结果汇总"

    # 填充总表内容
    fill_worksheet_content(ws, dataset, dataset_data, is_summary=True)

    wb.save(excel_path)
    print(f"✅ [总表] 生成完成：{excel_path}")

def generate_single_repeat_excel(dataset, dataset_data, repeat_idx, output_dir):
    """生成单次重复实验独立表"""
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, f"repeat_{repeat_idx + 1}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = f"第{repeat_idx + 1}次重复"

    # 填充单次表内容
    fill_worksheet_content(ws, dataset, dataset_data, is_summary=False, repeat_idx=repeat_idx)

    wb.save(excel_path)

def main():
    print("=" * 75)
    print("🚀 重构版 | 固定布局总表 + 10次重复实验独立表")
    print("=" * 75)

    if not os.path.exists(ROOT_DIR):
        print(f"❌ 根目录不存在：{ROOT_DIR}")
        return

    # 1. 全量读取所有数据
    all_data = load_all_experiment_data()

    # 2. 逐数据集处理
    for dataset in TARGET_DATASETS:
        out_dir = os.path.join(OUTPUT_PARENT, dataset)
        print(f"\n📂 正在处理数据集：{dataset}")

        # 生成总表
        generate_dataset_summary_excel(dataset, all_data[dataset], out_dir)

        # 生成10次重复实验独立表
        print(f"  📄 正在生成 10 次重复实验单独表...")
        for i in range(10):
            generate_single_repeat_excel(dataset, all_data[dataset], i, out_dir)
        print(f"  ✅ 10 次重复实验单独表全部生成完成！")

    print(f"\n🎉 所有任务执行完成！")

if __name__ == "__main__":
    main()