#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
完美增强版 ✅
✅ 原有总表逻辑 100% 保留（独立表头、分段、高亮）
✅ 新增：为 10 次重复实验单独生成 Excel 表
✅ 单独表：不计算均值，直接展示单次原始数据
✅ 单独表结构：和总表完全一致（独立表头、分段）
✅ 保存路径：{dataset}/repeat_{i}.xlsx
"""

import os
import json
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ======================== 配置 ========================
ROOT_DIR = r"D:\user\Documents\ai\paper\1_process\outputs\result"
OUTPUT_PARENT = r"D:\user\Documents\ai\paper\1_process\outputs\result"

MODEL_LIST = ["darm", "groupdro", "erm", "mixstyle", "mldg", "vrex", "irm", "dpjdg", "mcpdg"]
TARGET_DATASETS = ["phm", "pu"]
TARGET_TASKS = {"phm": ["T1", "T2", "T3", "T4"], "pu": ["T5", "T6", "T7", "T8"]}
TARGET_DATA_SIZES = ["1", "5", "10"]
METRIC_MAP = {"acc": "Acc", "precision_macro": "Pre", "recall_macro": "Rec", "f1_macro": "F1"}

# 总表高亮样式
HIGHLIGHT_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
# 列宽设置
COL_WIDTH_MODEL = 14
COL_WIDTH_DEFAULT = 18


# ======================================================

def get_mean_value(cell_text):
    """提取均值用于比较，空值返回极小值"""
    if cell_text == "-" or cell_text is None or cell_text == "":
        return -9999
    try:
        return float(cell_text.split("±")[0])
    except:
        return -9999


def load_single_json(json_path):
    """读取单个JSON指标文件"""
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {k: round(data.get(k, 0) * 100, 4) for k in METRIC_MAP}
    except Exception as e:
        # print(f"⚠️ 读取失败：{json_path} | 错误：{str(e)}") # 可选择关闭单次读取警告
        return None


def load_all_experiment_data():
    """
    全量读取所有实验数据
    返回结构：
    {
        "dataset": {
            "size": {
                "task": {
                    "model": [
                        {"acc": ..., "precision_macro": ...}, # 第1次重复
                        {"acc": ..., "precision_macro": ...}, # 第2次重复
                        ...
                    ]
                }
            }
        }
    }
    """
    full_data = {
        ds: {
            sz: {tk: {m: [] for m in MODEL_LIST} for tk in TARGET_TASKS[ds]}
            for sz in TARGET_DATA_SIZES
        }
        for ds in TARGET_DATASETS
    }

    for model in MODEL_LIST:
        model_path = os.path.join(ROOT_DIR, model)
        if not os.path.exists(model_path):
            continue

        for folder in os.listdir(model_path):
            if "-" not in folder:
                continue
            base_part, repeat_str = folder.rsplit("-", 1)
            if not repeat_str.isdigit():
                continue
            repeat_idx = int(repeat_str) - 1  # 转换为 0-9 的索引

            parts = base_part.split("_")
            if len(parts) != 4:
                continue
            _, dataset, task, size = parts

            if dataset not in TARGET_DATASETS or task not in TARGET_TASKS[dataset] or size not in TARGET_DATA_SIZES:
                continue

            json_file = os.path.join(model_path, folder, "final_test_metrics.json")
            if os.path.exists(json_file):
                result = load_single_json(json_file)
                if result:
                    # 确保列表长度足够，按索引插入
                    while len(full_data[dataset][size][task][model]) <= repeat_idx:
                        full_data[dataset][size][task][model].append(None)
                    full_data[dataset][size][task][model][repeat_idx] = result

    print("✅ 所有实验数据已全部读取完成")
    return full_data


def calc_mean_std(metric_list):
    """计算均值±标准差，无数据返回"-" """
    result = {}
    for k in METRIC_MAP:
        values = [m[k] for m in metric_list if m is not None]
        if not values:
            result[k] = "-"
        else:
            mean = np.mean(values)
            std = np.std(values)
            result[k] = f"{mean:.2f}±{std:.2f}"
    return result


# ======================== 原有总表生成逻辑（完全未改动） ========================
def generate_dataset_summary_excel(dataset, dataset_data, output_dir):
    """【原有功能】生成带均值±标准差和高亮的总表"""
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, f"{dataset}_result.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "结果汇总"

    current_row = 1
    model_count = len(MODEL_LIST)

    for size in TARGET_DATA_SIZES:
        header = ["模型"]
        for task in TARGET_TASKS[dataset]:
            for metric_short in METRIC_MAP.values():
                header.append(f"{task}_{size}_{metric_short}")

        for col_idx, header_text in enumerate(header, 1):
            ws.cell(row=current_row, column=col_idx, value=header_text)
        data_start_row = current_row + 1
        data_end_row = current_row + model_count

        for model_idx, model in enumerate(MODEL_LIST, data_start_row):
            ws.cell(row=model_idx, column=1, value=model)
            col_idx = 2
            for task in TARGET_TASKS[dataset]:
                metrics = calc_mean_std(dataset_data[size][task][model])
                for metric_key in METRIC_MAP:
                    ws.cell(row=model_idx, column=col_idx, value=metrics[metric_key])
                    col_idx += 1

        max_col = len(header)
        for col in range(2, max_col + 1):
            values = []
            for r in range(data_start_row, data_end_row + 1):
                cell_val = ws.cell(row=r, column=col).value
                values.append(get_mean_value(cell_val))

            max_val = max(values)
            if max_val == -9999:
                continue

            for r in range(data_start_row, data_end_row + 1):
                current_val = get_mean_value(ws.cell(row=r, column=col).value)
                if current_val == max_val:
                    ws.cell(row=r, column=col).fill = HIGHLIGHT_FILL

        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = COL_WIDTH_MODEL if col == 1 else COL_WIDTH_DEFAULT

        current_row = data_end_row + 2

    wb.save(excel_path)
    print(f"✅ [总表] 生成完成：{excel_path}")


# ======================== 新增：单次重复实验表生成 ========================
def generate_single_repeat_excel(dataset, dataset_data, repeat_idx, output_dir):
    """【新增功能】为第 repeat_idx+1 次重复实验生成单独表（无均值，原始数据）"""
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, f"repeat_{repeat_idx + 1}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = f"第{repeat_idx + 1}次重复"

    current_row = 1
    model_count = len(MODEL_LIST)

    for size in TARGET_DATA_SIZES:
        # 表头和总表完全一致
        header = ["模型"]
        for task in TARGET_TASKS[dataset]:
            for metric_short in METRIC_MAP.values():
                header.append(f"{task}_{size}_{metric_short}")

        for col_idx, header_text in enumerate(header, 1):
            ws.cell(row=current_row, column=col_idx, value=header_text)
        data_start_row = current_row + 1
        data_end_row = current_row + model_count

        for model_idx, model in enumerate(MODEL_LIST, data_start_row):
            ws.cell(row=model_idx, column=1, value=model)
            col_idx = 2
            for task in TARGET_TASKS[dataset]:
                # 直接取第 repeat_idx 次的原始数据，不计算均值
                repeat_data_list = dataset_data[size][task][model]
                val = "-"
                if repeat_idx < len(repeat_data_list) and repeat_data_list[repeat_idx] is not None:
                    data = repeat_data_list[repeat_idx]
                    for metric_key in METRIC_MAP:
                        # 原始数据保留两位小数
                        val = f"{data[metric_key]:.2f}"
                        ws.cell(row=model_idx, column=col_idx, value=val)
                        col_idx += 1
                else:
                    # 无数据则填 "-"
                    for _ in METRIC_MAP:
                        ws.cell(row=model_idx, column=col_idx, value="-")
                        col_idx += 1

        # 设置列宽
        max_col = len(header)
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = COL_WIDTH_MODEL if col == 1 else COL_WIDTH_DEFAULT

        # 空行分隔
        current_row = data_end_row + 2

    wb.save(excel_path)
    # print(f"  ✔️ [单次表] 生成完成：repeat_{repeat_idx + 1}.xlsx") # 可选择开启详细打印


def main():
    print("=" * 75)
    print("🚀 完美增强版 | 保留总表 + 新增10次重复实验单独表")
    print("=" * 75)

    if not os.path.exists(ROOT_DIR):
        print(f"❌ 根目录不存在：{ROOT_DIR}")
        return

    # 1. 全量读取所有数据
    all_data = load_all_experiment_data()

    # 2. 逐数据集处理
    for dataset in TARGET_DATASETS:
        out_dir = os.path.join(OUTPUT_PARENT, dataset)
        print(f"\n📂 正在处理数据集：{dataset}")

        # 【原有功能 1/2】生成带均值和高亮的总表（逻辑完全未动）
        generate_dataset_summary_excel(dataset, all_data[dataset], out_dir)

        # 【新增功能 2/2】为 10 次重复实验单独生成表
        print(f"  📄 正在生成 10 次重复实验单独表...")
        for i in range(10):
            generate_single_repeat_excel(dataset, all_data[dataset], i, out_dir)
        print(f"  ✅ 10 次重复实验单独表全部生成完成！")

    print("\n🎉 全部任务完美完成！")


if __name__ == "__main__":
    main()