import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_points

# ===================== 配置项（务必修改！）=====================
FOLDER_PATH = r"D:\user\code\AI\其他\成绩脚本\插入图片\全部"  # 示例：r"C:\Users\XXX\Pictures\旅游照片"
# 生成的Excel保存路径（默认保存在当前目录，可自定义）
EXCEL_SAVE_PATH = r"图片汇总表2.xlsx"
SUPPORTED_FORMATS = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp')
IMG_DISPLAY_WIDTH = 120  # 图片显示宽度（像素）

def export_images_to_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "图片汇总"

    # 设置表头和列宽（加宽B列，避免图片被遮挡）
    ws['A1'] = "图片名称（无后缀）"
    ws['B1'] = "图片预览"
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 80  # 加宽B列，确保图片不被遮挡

    row_num = 2
    # 遍历图片（兼容中文路径）
    for filename in os.listdir(FOLDER_PATH):
        if filename.lower().endswith(SUPPORTED_FORMATS):
            try:
                # 1. 拼接绝对路径（强制转成规范路径，避免中文/空格问题）
                img_path = os.path.abspath(os.path.join(FOLDER_PATH, filename))
                if not os.path.exists(img_path):
                    print(f"❌ 图片不存在：{img_path}")
                    continue

                # 2. 写入文件名
                file_name_no_ext = os.path.splitext(filename)[0]
                ws[f'A{row_num}'] = file_name_no_ext

                # 3. 加载图片（关键：指定格式，避免pillow解析失败）
                img = Image(img_path)
                # 缩放图片（保持宽高比）
                scale_ratio = IMG_DISPLAY_WIDTH / img.width
                img.width = IMG_DISPLAY_WIDTH
                img.height = int(img.height * scale_ratio)

                # 4. 锚定图片到单元格（兼容所有openpyxl版本的写法）
                # 锚点格式：B{row_num}（直接指定单元格，而非手动计算索引）
                img.anchor = f'B{row_num}'

                # 5. 添加图片到工作表
                ws.add_image(img)

                # 6. 调整行高（像素转磅，精准适配）
                row_height = pixels_to_points(img.height) + 15
                ws.row_dimensions[row_num].height = row_height

                print(f"✅ 成功处理：{filename}")
                row_num += 1

            except Exception as e:
                print(f"❌ 处理{filename}失败：{str(e)}")

    # 保存文件（强制覆盖，避免权限问题）
    wb.save(EXCEL_SAVE_PATH)
    print(f"\n🎉 生成完成！Excel路径：{EXCEL_SAVE_PATH}")
    print(f"📊 成功处理 {row_num - 2} 张图片")

if __name__ == "__main__":
    if not os.path.exists(FOLDER_PATH):
        print(f"❌ 文件夹不存在：{FOLDER_PATH}")
    else:
        # 检查依赖版本（关键！）
        try:
            import openpyxl
            import PIL
            print(f"✅ 依赖版本：openpyxl={openpyxl.__version__}, PIL={PIL.__version__}")
        except ImportError as e:
            print(f"❌ 依赖缺失：{e}")
            exit(1)
        export_images_to_excel()